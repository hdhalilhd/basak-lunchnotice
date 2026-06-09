import os
import datetime as dt
import requests
from openpyxl import load_workbook

# === AYARLAR ===
CHAT_ID = int(os.getenv("TELEGRAM_CHAT_ID", "1677402217"))  # Sana gelecek özel mesaj
GROUP_CHAT_ID = int(os.getenv("GROUP_CHAT_ID", "0"))        # BAŞAK YEMEK MENÜ grubuna gidecek mesaj

TOKEN = os.getenv("TELEGRAM_BOT_TOKEN", "")
EXCEL_PATH = os.getenv("EXCEL_PATH", "BASAK_Yemek_Listesi.xlsx")
SHEET_NAME = os.getenv("SHEET_NAME", "")
SEND_TOMORROW = os.getenv("SEND_TOMORROW", "false").lower() == "true"
# =================

MONTHS = {
    "Ocak": 1, "Şubat": 2, "Mart": 3, "Nisan": 4, "Mayıs": 5, "Haziran": 6,
    "Temmuz": 7, "Ağustos": 8, "Eylül": 9, "Ekim": 10, "Kasım": 11, "Aralık": 12
}

def parse_tr_date(s):
    if not s:
        return None
        
    # Excel hücreyi tarih objesi olarak verirse doğrudan al
    if isinstance(s, (dt.datetime, dt.date)):
        return s.date() if isinstance(s, dt.datetime) else s

    # Metin olarak girilmişse parçalayarak al
    parts = str(s).strip().split()
    if len(parts) != 3:
        return None
    try:
        return dt.date(int(parts[2]), MONTHS.get(parts[1]), int(parts[0]))
    except:
        return None

def get_menu(target_date):
    try:
        wb = load_workbook(EXCEL_PATH, data_only=True)
        if SHEET_NAME:
            ws = wb[SHEET_NAME]
        else:
            ws = wb.active
    except Exception as e:
        return f"❗ Excel dosyası veya sayfası bulunamadı: {e}"

    for row in ws.iter_rows(min_row=2, values_only=True):
        # Satır tamamen boşsa atla (Çalışan koddan ilham alındı)
        if not any(row):
            continue

        # Gelen satırı listeye çevirip ilk 6 elemanını al ve eksikleri tamamla
        veri = list(row)[:6]
        veri += [None] * (6 - len(veri))

        t, gun, corba, ana, yard, tatli = veri

        d = parse_tr_date(t)
        if d == target_date:
            label = "Yarın" if SEND_TOMORROW else "Bugün"
            return (
                f"🍽️ BAŞAK TRAKTÖR\n"
                f"📌 {gun} {label} Yemek Menüsü\n"
                f"🍲 Çorba: {corba}\n"
                f"🍝 Ana Yemek: {ana}\n"
                f"🍖 Yardımcı: {yard}\n"
                f"🍮 Tatlı/Meyve: {tatli}"
            )

    return f"❗ Bugün için yemek menüsü bulunamadı."

def send_telegram(msg, chat_id):
    if not TOKEN:
        print("HATA: TELEGRAM_BOT_TOKEN bulunamadı.")
        return
        
    url = f"https://api.telegram.org/bot{TOKEN}/sendMessage"
    try:
        # Timeout ve hata yönetimi eklendi
        r = requests.post(url, json={"chat_id": chat_id, "text": msg}, timeout=20)
        r.raise_for_status()
        print(f"Mesaj başarıyla gönderildi (ID: {chat_id})")
    except Exception as e:
        print(f"Mesaj gönderilirken hata oluştu (ID: {chat_id}): {e}")

if __name__ == "__main__":
    # GitHub Actions üzerinde Türkiye saatini (UTC+3) zorunlu kıl
    turkey_tz = dt.timezone(dt.timedelta(hours=3))
    today = dt.datetime.now(turkey_tz).date()
    
    target = today + dt.timedelta(days=1) if SEND_TOMORROW else today

    # Haftasonu: Cumartesi=5, Pazar=6 -> mesaj yok
    if target.weekday() >= 5:
        print("Hafta sonu - mesaj gönderilmedi.")
        raise SystemExit(0)

    msg = get_menu(target)
    
    # 1. Admin'e her durumda mesaj gönder (Hata mesajları dahil)
    send_telegram(msg, CHAT_ID)
    
    # 2. Gruba sadece menü başarıyla bulunduysa ve ID 0'dan farklıysa gönder
    if not msg.startswith("❗") and GROUP_CHAT_ID != 0:
        send_telegram(msg, GROUP_CHAT_ID)
    else:
        print("Hata, tatil günü veya grup ID'si 0 olduğu için gruba mesaj atılmadı.")
